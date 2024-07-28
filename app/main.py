import discord
from discord import app_commands
import os
from dotenv import load_dotenv
import openpyxl
from openpyxl import Workbook
import uuid

# .envファイルの内容を読み込見込む
load_dotenv()

TOKEN = os.getenv("TOKEN")
intents = discord.Intents.all()
client = discord.Client(intents=intents)
tree = app_commands.CommandTree(client)

# Excelファイルの名前
MAIN_FILE = 'Denpa_ArenaData.xlsx'
WINNER_FILE = 'Denpa_ArenaData_Winner.xlsx'

@client.event
async def on_ready():
    print(f'We have logged in as {client.user}')
    try:
        synced = await tree.sync()
        print(f"Synced {len(synced)} command(s)")
    except Exception as e:
        print(e)
    client.add_view(ChannelCreationButton())
    # Excelファイルが存在しない場合は新規作成
    try:
        mainFILE = openpyxl.load_workbook(MAIN_FILE)
        print('Main File Loaded')
    except FileNotFoundError:
        mainFILE = Workbook()
        sheet = mainFILE.active
        sheet['A1'] = 'Name'
        sheet['B1'] = 'Level'
        sheet['C1'] = 'Odds'
        sheet['D1'] = 'ID'
        mainFILE.save(MAIN_FILE)
        mainFILE.save(WINNER_FILE)
        print('Main File Created')
    try:
        winnerFILE = openpyxl.load_workbook(WINNER_FILE)
        print('Winner File Loaded')
    except FileNotFoundError:
        winnerFILE = Workbook()
        sheet = winnerFILE.active
        sheet['A1'] = 'ID'
        sheet['B1'] = 'Winner'
        winnerFILE.save(WINNER_FILE)
        print('Winner File Created')


class ChannelCreationButton(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)  # タイムアウトをNoneに設定

    @discord.ui.button(label='チャンネルを作成', style=discord.ButtonStyle.primary, custom_id='create_channel')
    async def create_channel(self, interaction: discord.Interaction, button: discord.ui.Button):
        await add(interaction)

@tree.command(name='show_button', description='チャンネル作成ボタンを表示')
async def show_button(interaction: discord.Interaction):
    view = ChannelCreationButton()
    await interaction.response.send_message("以下のボタンをクリックしてプライベートチャンネルを作成できます：", view=view)

async def create_channel(interaction: discord.Interaction):
    # チャンネル名を設定（ユーザー名をベースにする）
    channel_name = f"private-{interaction.user.name.lower()}"

    # 既存のチャンネルをチェック
    existing_channel = discord.utils.get(interaction.guild.channels, name=channel_name)
    if existing_channel:
        await interaction.response.send_message(f'チャンネル {existing_channel.mention} は既に存在します。', ephemeral=True)
        return

    # オーバーライド設定を作成
    overwrites = {
        interaction.guild.default_role: discord.PermissionOverwrite(read_messages=False, send_messages=False),
        interaction.guild.me: discord.PermissionOverwrite(read_messages=True, send_messages=True),
        interaction.user: discord.PermissionOverwrite(read_messages=True, send_messages=True)
    }
    # 管理者ロールを持つメンバーに権限を付与
    for role in interaction.guild.roles:
        if role.permissions.administrator:
            overwrites[role] = discord.PermissionOverwrite(read_messages=True, send_messages=True)

    # 新しいチャンネルを作成
    channel = await interaction.guild.create_text_channel(channel_name, overwrites=overwrites)

    await interaction.response.send_message(f'プライベートチャンネル {channel.mention} を作成しました。', ephemeral=True)

    # 作成したチャンネルにメッセージを送信
    await channel.send(f'このチャンネルは作成者（あなた）と管理者のみが閲覧できます。\nデータの追加の方法やAIの使用方法などは使用方法チャンネルをご覧ください')

async def button(interaction: discord.Interaction):
    button = discord.ui.Button(label='Adding Data', style=discord.ButtonStyle.primary)
    view = discord.ui.View()
    view.add_item(button)
    await interaction.response.send_message(view=view, ephemeral=True)


# Pingを返す
@tree.command(name='ping', description='ping')
async def ping(interaction: discord.Interaction):
    latency = round(client.latency * 1000)
    await interaction.response.send_message(f'Your ping is {latency}ms')


@client.event
async def on_message(message):
    if message.author == client.user:
        return

    if message.content == 'hello':
        await message.channel.send('hay')


client.run(TOKEN)